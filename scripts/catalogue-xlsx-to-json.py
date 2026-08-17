#!/usr/bin/env python3
"""Dump the Catalogue sheet of a filled catalogue workbook to JSON.

This script is deliberately DUMB. It reads cells and writes JSON — nothing
else. Every domain rule (which categories exist, what a valid subcategory is,
whether MRP may sit below the selling price) is enforced by
scripts/import-catalogue.ts, which can import the real `MASTER_CATEGORY_TREE`
from src/constants.ts. Re-stating those rules here in Python would create a
second copy that drifts the first time a category is added.

Python is used at all for one reason: reading .xlsx needs openpyxl, and adding
an Excel parser to package.json for a back-office import is not a trade this
repo makes (see the dependency note in src/admin/exportKitPdf.ts).

Columns are matched by HEADER NAME, not position, so the sheet can be
reordered and extra columns appended without touching this file. Computed
columns (Margin %, Stock Status, Stock per Log, Log Match) are skipped: their
cells hold formulas, and the values they compute are not inputs.

    python3 scripts/catalogue-xlsx-to-json.py <workbook.xlsx> [-o out.json]

Output: {"source": "...", "sheet": "Catalogue", "rows": [{header: value, ...}]}
with each row carrying `_row` (the spreadsheet row number) so the importer can
point at a real line when something is wrong.
"""
import argparse
import datetime as dt
import json
import sys

try:
    import openpyxl
except ImportError:                                          # pragma: no cover
    sys.exit('openpyxl is required:  pip3 install openpyxl')

# Cells here hold formulas; whatever they evaluate to is derived from the
# columns we DO read, so importing them would be importing our own arithmetic.
COMPUTED = {'Margin %', 'Stock Status', 'Stock per Log', 'Log Match'}

# The workbook ships one obviously-disposable demonstration row.
EXAMPLE_MARKER = 'EXAMPLE ROW'


def clean(value):
    """Excel cell -> JSON-safe value. Blank-ish cells become None."""
    if value is None:
        return None
    if isinstance(value, str):
        s = value.strip()
        # A formula string means data_only was off and the cell is computed;
        # the caller filters those columns out, but guard anyway.
        return None if not s or s.startswith('=') else s
    if isinstance(value, (dt.datetime, dt.date)):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, bool):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument('workbook')
    ap.add_argument('-o', '--out', default='catalogue-import.json')
    ap.add_argument('--sheet', default='Catalogue')
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.workbook, data_only=False, read_only=True)
    if args.sheet not in wb.sheetnames:
        sys.exit(f'No "{args.sheet}" sheet in {args.workbook}. Found: {", ".join(wb.sheetnames)}')
    ws = wb[args.sheet]

    rows = iter(ws.iter_rows(values_only=True))
    try:
        header = [clean(h) for h in next(rows)]
    except StopIteration:
        sys.exit(f'"{args.sheet}" is empty.')

    wanted = [(i, h) for i, h in enumerate(header) if h and h not in COMPUTED]
    if not any(h == 'Product ID' for _, h in wanted):
        sys.exit('The sheet has no "Product ID" column — is this the Catalogue sheet?')

    out, skipped_blank, skipped_example = [], 0, 0
    for n, raw in enumerate(rows, start=2):
        rec = {h: clean(raw[i]) if i < len(raw) else None for i, h in wanted}
        if not rec.get('Product ID'):
            skipped_blank += 1
            continue
        notes = rec.get('Notes') or ''
        if isinstance(notes, str) and notes.upper().startswith(EXAMPLE_MARKER):
            skipped_example += 1
            continue
        rec['_row'] = n
        out.append(rec)

    payload = {'source': args.workbook, 'sheet': args.sheet, 'rows': out}
    with open(args.out, 'w', encoding='utf-8') as fh:
        json.dump(payload, fh, ensure_ascii=False, indent=1)

    print(f'{args.workbook} [{args.sheet}] -> {args.out}')
    print(f'  {len(out)} product row(s), {len(wanted)} column(s) read')
    if skipped_example:
        print(f'  {skipped_example} example row(s) skipped')
    if skipped_blank:
        print(f'  {skipped_blank} blank row(s) skipped')
    if not out:
        print('  Nothing to import — every row was blank or an example row.')
        return 1
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
