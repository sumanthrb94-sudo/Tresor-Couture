"""Build the Tresor Couture master catalogue workbooks.

Two files:
  tresor-catalogue-TEMPLATE.xlsx  — empty, dropdowns + formulas + one example row
  tresor-catalogue-SAMPLE.xlsx    — the real 50-product catalogue with embedded
                                    photos and a reconciled month of stock log

Schema is derived from the production data model (src/types.ts Fabric) plus the
operational fields the code does not yet track (buying price, supplier, HSN,
reorder level, stock movements).
"""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

SCRATCH = '/tmp/claude-0/-home-user-Tresor-Couture/032c8341-8752-527a-934b-e8820f30a5ea/scratchpad'
CAT = json.load(open(f'{SCRATCH}/catalogue.json'))
CAT.sort(key=lambda p: (p['master'], p['sub'], p['id']))

ARIAL = 'Arial'
HDR_FILL = PatternFill('solid', fgColor='1F2A44')
HDR_FONT = Font(name=ARIAL, size=9, bold=True, color='FFFFFF')
BASE = Font(name=ARIAL, size=10)
BLUE = Font(name=ARIAL, size=10, color='0000FF')       # hardcoded inputs
GREY_FILL = PatternFill('solid', fgColor='F2F2F2')
THIN = Border(bottom=Side(style='thin', color='D9D9D9'))

# ── Column model for the Catalogue sheet ─────────────────────────────────────
COLS = [
    ('Photo', 13), ('Product ID', 16), ('Supplier Code', 13), ('Product Name', 34),
    ('Brand', 14), ('Master Category', 16), ('Sub Category / Design', 18), ('Material', 12),
    ('Unit Type', 11), ('Bundle (m)', 9), ('Buying Price (₹)', 13), ('Selling Price (₹)', 13),
    ('MRP (₹)', 11), ('Margin %', 9), ('GST Rate', 9), ('HSN Code', 10),
    ('Stock Qty', 9), ('Reorder Lvl', 9), ('Stock Status', 13), ('Stock per Log', 11),
    ('Log Match', 9), ('First Received', 12), ('Supplier', 16), ('Image URL', 40),
    ('Photo Quality', 13), ('Live on Site', 10), ('Listing Status', 12), ('Sticker', 11),
    ('Notes', 40),
]
C = {name: get_column_letter(i + 1) for i, (name, _) in enumerate(COLS)}

LISTS = {
    'Master Category': ['Fabrics', 'Dyeable Fabrics', 'Laces', 'Sarees', 'Lehenga Cholis',
                        'Anarkalis', 'Western Wear', 'Studios Prêt'],
    'Material': ['Silk', 'Silk blend', 'Cotton', 'Wool', 'Linen', 'Satin', 'Net',
                 'Georgette', 'Organza', 'Mixed'],
    'Unit Type': ['unit', 'per meter', 'bundle'],
    'GST Rate': [0, 0.05, 0.12, 0.18],
    'Photo Quality': ['Studio', 'Acceptable', 'Needs reshoot'],
    'Live on Site': ['Yes', 'No'],
    'Listing Status': ['Active', 'Draft', 'Retired'],
    'Sticker': ['New In', 'Bestseller', 'Trending', 'Limited'],
    'Movement': ['Opening stock', 'Received (new stock)', 'Sold - Online', 'Sold - In Store',
                 'Sold - Export', 'Returned (back to stock)', 'Damaged / Write-off',
                 'Sample / Gift', 'Count correction'],
    'Supplier': ['Supplier A (rename me)', 'Supplier B (rename me)', 'Atelier — own production'],
    'Channel': ['Website - COD', 'Website - Prepaid', 'In Store', 'WhatsApp / DM', 'Export'],
    'Payment': ['COD - to collect', 'COD - collected', 'Prepaid - paid', 'Refunded', 'Partly refunded'],
    'Fulfilment': ['Placed', 'Packed', 'Shipped', 'Delivered', 'Cancelled'],
    'Return Status': ['Requested', 'Approved', 'Pickup arranged', 'Received back', 'Refunded', 'Replaced', 'Rejected'],
    'Task Area': ['Marketing', 'SEO', 'Compliance', 'Finance', 'Catalogue', 'Security', 'Operations'],
    'Task Status': ['Pending', 'In progress', 'Blocked', 'Done'],
}

MAX_ROWS = 200   # validations + prefilled formulas reach down to here
LOG_ROWS = 600


def build(path: str, with_data: bool) -> None:
    wb = Workbook()

    # ── Lists sheet (dropdown sources) ───────────────────────────────────────
    ws_l = wb.active
    ws_l.title = 'Lists'
    list_ranges = {}
    for col_i, (name, values) in enumerate(LISTS.items(), start=1):
        letter = get_column_letter(col_i)
        ws_l[f'{letter}1'] = name
        ws_l[f'{letter}1'].font = HDR_FONT
        ws_l[f'{letter}1'].fill = HDR_FILL
        for row_i, v in enumerate(values, start=2):
            cell = ws_l[f'{letter}{row_i}']
            cell.value = v
            cell.font = BASE
            if name == 'GST Rate':
                cell.number_format = '0%'
        ws_l.column_dimensions[letter].width = 22
        list_ranges[name] = f'Lists!${letter}$2:${letter}${len(values) + 1}'

    # ── Catalogue sheet ──────────────────────────────────────────────────────
    ws = wb.create_sheet('Catalogue')
    for i, (name, width) in enumerate(COLS, start=1):
        cell = ws.cell(row=1, column=i, value=name)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = 'E2'
    ws.auto_filter.ref = f'A1:AC{MAX_ROWS}'

    def formulas_for(r: int) -> dict:
        return {
            'Margin %': f'=IF(OR({C["Buying Price (₹)"]}{r}="",{C["Selling Price (₹)"]}{r}=""),"",'
                        f'({C["Selling Price (₹)"]}{r}-{C["Buying Price (₹)"]}{r})/{C["Selling Price (₹)"]}{r})',
            'Stock Status': f'=IF({C["Stock Qty"]}{r}="","",IF({C["Stock Qty"]}{r}<=0,"OUT OF STOCK",'
                            f'IF({C["Stock Qty"]}{r}<={C["Reorder Lvl"]}{r},"REORDER","OK")))',
            'Stock per Log': f'=IF({C["Product ID"]}{r}="","",'
                             f"SUMIFS('Stock Log'!$D$2:$D$600,'Stock Log'!$B$2:$B$600,{C['Product ID']}{r}))",
            'Log Match': f'=IF({C["Product ID"]}{r}="","",'
                         f'IF({C["Stock per Log"]}{r}={C["Stock Qty"]}{r},"OK","CHECK"))',
        }

    def style_row(r: int) -> None:
        for i in range(1, len(COLS) + 1):
            cell = ws.cell(row=r, column=i)
            if cell.font.name != ARIAL:
                cell.font = BASE
            cell.border = THIN
        for name, fmt in [('Buying Price (₹)', '"₹"#,##0'), ('Selling Price (₹)', '"₹"#,##0'),
                          ('MRP (₹)', '"₹"#,##0'), ('Margin %', '0.0%'), ('GST Rate', '0%'),
                          ('First Received', 'yyyy-mm-dd')]:
            ws[f'{C[name]}{r}'].number_format = fmt

    # illustrative buying prices (BLUE input; documented in How To Use)
    def est_cost(p: dict) -> int:
        share = 0.62 if p['master'] in ('Lehenga Cholis', 'Sarees') else 0.58
        return int(round(p['price'] * share, -1))

    log_entries = []   # (date, sku, movement, qty, ref, note)
    if with_data:
        SOLD = {'6343': ('2026-08-02', 'ORD-8H2K'), '6988': ('2026-08-04', 'ORD-8J7Q'),
                'HA4536': ('2026-08-07', 'ORD-8N1D'), 'HA6320': ('2026-08-09', 'ORD-8Q5S'),
                'HA3858': ('2026-08-11', 'ORD-8T3M'), 'MI263': ('2026-08-13', 'ORD-8V9A')}
        for p in CAT:
            received = '2026-07-31' if p['master'] in ('Lehenga Cholis', 'Sarees') else '2026-07-09'
            if p['master'] in ('Lehenga Cholis', 'Sarees'):
                log_entries.append((received, p['id'], 'Received (new stock)', p['stock'], 'PO-DROP-01', 'Bridal drop intake'))
                if p['id'] == 'lc-zf-rose':
                    log_entries.append(('2026-08-05', p['id'], 'Sold - Online', -1, 'ORD-8L4R', 'Prepaid order'))
                    log_entries.append(('2026-08-12', p['id'], 'Returned (back to stock)', 1, 'RET-8L4R', 'Size exchange requested; restocked after inspection'))
            else:
                sold = 1 if p['id'] in SOLD else 0
                adj = -1 if p['id'] == 'RI280' else 0
                opening = p['stock'] + sold - adj
                log_entries.append((received, p['id'], 'Opening stock', opening, 'IMPORT-0709', 'Stock on hand when this sheet was started'))
                if sold:
                    d, ref = SOLD[p['id']]
                    log_entries.append((d, p['id'], 'Sold - Online', -1, ref, 'COD order'))
                if adj:
                    log_entries.append(('2026-08-02', p['id'], 'Damaged / Write-off', -1, 'ADJ-001', '1 reel water-damaged'))

        for r, p in enumerate(CAT, start=2):
            reorder = 2 if p['master'] in ('Lehenga Cholis', 'Sarees') else 1
            needs_reshoot = p['id'] in ('HA6378', 'MI263', 'RI5687')
            vals = {
                'Product ID': p['id'], 'Supplier Code': p['code'], 'Product Name': p['name'],
                'Brand': p['brand'], 'Master Category': p['master'], 'Sub Category / Design': p['sub'],
                'Material': p['material'], 'Unit Type': p['unit'], 'Bundle (m)': p['bundle'],
                'Buying Price (₹)': est_cost(p), 'Selling Price (₹)': p['price'], 'MRP (₹)': p['mrp'],
                'GST Rate': 0.05, 'HSN Code': None, 'Stock Qty': p['stock'], 'Reorder Lvl': reorder,
                'First Received': '2026-07-31' if p['master'] in ('Lehenga Cholis', 'Sarees') else '2026-07-09',
                'Supplier': 'Atelier — own production' if p['master'] in ('Lehenga Cholis', 'Sarees') else 'Supplier A (rename me)',
                'Image URL': f"https://tresorcouture.in{p['photo']}" if p['photo'].startswith('/') else '',
                'Photo Quality': 'Needs reshoot' if needs_reshoot else 'Studio',
                'Live on Site': 'Yes', 'Listing Status': 'Active', 'Sticker': p['sticker'] or None,
                'Notes': 'Reshoot flagged — warehouse snapshot' if needs_reshoot else None,
            }
            for name, v in vals.items():
                cell = ws[f'{C[name]}{r}']
                cell.value = v
                if name == 'Buying Price (₹)':
                    cell.font = BLUE
            for name, f in formulas_for(r).items():
                ws[f'{C[name]}{r}'] = f
            style_row(r)
            ws.row_dimensions[r].height = 56
            try:
                img = XLImage(f"{SCRATCH}/thumbs/{p['id']}.jpg")
                img.anchor = f'A{r}'
                ws.add_image(img)
            except FileNotFoundError:
                pass
        data_rows = len(CAT) + 1
    else:
        # one example row, clearly disposable
        r = 2
        example = {
            'Product ID': 'lc-example-1', 'Supplier Code': 'HA0000', 'Product Name': 'Example Lace Border · Gold',
            'Brand': 'TRESOR COUTURE', 'Master Category': 'Laces', 'Sub Category / Design': 'Trim & Edging',
            'Material': 'Net', 'Unit Type': 'bundle', 'Bundle (m)': 9, 'Buying Price (₹)': 700,
            'Selling Price (₹)': 1299, 'MRP (₹)': 1299, 'GST Rate': 0.05, 'HSN Code': None,
            'Stock Qty': 3, 'Reorder Lvl': 1, 'First Received': '2026-08-01',
            'Supplier': 'Supplier A (rename me)', 'Image URL': 'https://tresorcouture.in/products/lace/HA0000.jpg',
            'Photo Quality': 'Studio', 'Live on Site': 'Yes', 'Listing Status': 'Active',
            'Sticker': 'New In', 'Notes': 'EXAMPLE ROW — replace with your first product',
        }
        for name, v in example.items():
            cell = ws[f'{C[name]}{r}']
            cell.value = v
            if name == 'Buying Price (₹)':
                cell.font = BLUE
        for name, f in formulas_for(r).items():
            ws[f'{C[name]}{r}'] = f
        style_row(r)
        for i in range(1, len(COLS) + 1):
            ws.cell(row=r, column=i).fill = GREY_FILL
        data_rows = 2

    # prefill guarded formulas so new rows compute automatically
    for r in range(data_rows + 1, MAX_ROWS + 1):
        for name, f in formulas_for(r).items():
            ws[f'{C[name]}{r}'] = f
            ws[f'{C[name]}{r}'].font = BASE
        ws[f'{C["Margin %"]}{r}'].number_format = '0.0%'
        ws[f'{C["GST Rate"]}{r}'].number_format = '0%'

    # dropdown validations
    for name in ['Master Category', 'Material', 'Unit Type', 'GST Rate', 'Photo Quality',
                 'Live on Site', 'Listing Status', 'Sticker', 'Supplier']:
        dv = DataValidation(type='list', formula1=list_ranges[name], allow_blank=True,
                            showErrorMessage=True, errorTitle='Pick from the list',
                            error=f'Choose a value from the {name} dropdown (or edit the Lists sheet to add one).')
        ws.add_data_validation(dv)
        dv.add(f'{C[name]}2:{C[name]}{MAX_ROWS}')

    # conditional formatting on Stock Status
    red = PatternFill('solid', fgColor='F8D3D0')
    amber = PatternFill('solid', fgColor='FBEAC9')
    col = C['Stock Status']
    ws.conditional_formatting.add(f'{col}2:{col}{MAX_ROWS}',
        FormulaRule(formula=[f'${col}2="OUT OF STOCK"'], fill=red))
    ws.conditional_formatting.add(f'{col}2:{col}{MAX_ROWS}',
        FormulaRule(formula=[f'${col}2="REORDER"'], fill=amber))
    colm = C['Log Match']
    ws.conditional_formatting.add(f'{colm}2:{colm}{MAX_ROWS}',
        FormulaRule(formula=[f'${colm}2="CHECK"'], fill=red))

    # ── Stock Log sheet ──────────────────────────────────────────────────────
    ws_s = wb.create_sheet('Stock Log')
    log_cols = [('Date', 12), ('Product ID', 16), ('What Happened', 22), ('Qty  (+ in / − out)', 14),
                ('Stock After', 11), ('Reference (PO / Order no.)', 22), ('Notes', 44)]
    for i, (name, width) in enumerate(log_cols, start=1):
        cell = ws_s.cell(row=1, column=i, value=name)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        ws_s.column_dimensions[get_column_letter(i)].width = width
    ws_s.freeze_panes = 'A2'

    entries = sorted(log_entries) if with_data else [
        ('2026-08-01', 'lc-example-1', 'Received (new stock)', 3, 'PO-001', 'EXAMPLE ROW — first stock intake'),
    ]
    for r, (d, sku, mv, qty, ref, note) in enumerate(entries, start=2):
        balance = f'=IF($B{r}="","",SUMIFS($D$2:$D{r},$B$2:$B{r},$B{r}))'
        for i, v in enumerate([d, sku, mv, qty, balance, ref, note], start=1):
            cell = ws_s.cell(row=r, column=i, value=v)
            cell.font = BASE
            cell.border = THIN
        ws_s.cell(row=r, column=1).number_format = 'yyyy-mm-dd'
        if not with_data:
            for i in range(1, 7):
                ws_s.cell(row=r, column=i).fill = GREY_FILL

    for r in range(len(entries) + 2, LOG_ROWS + 1):
        ws_s[f'E{r}'] = f'=IF($B{r}="","",SUMIFS($D$2:$D{r},$B$2:$B{r},$B{r}))'
        ws_s[f'E{r}'].font = BASE
    ws_s.auto_filter.ref = f'A1:G{LOG_ROWS}'

    dv_mv = DataValidation(type='list', formula1=list_ranges['Movement'], allow_blank=True)
    ws_s.add_data_validation(dv_mv)
    dv_mv.add(f'C2:C{LOG_ROWS}')


    # ── Orders & Returns sheet ───────────────────────────────────────────────
    ws_o = wb.create_sheet('Orders & Returns')
    o_cols = [('Order Date', 12), ('Order No.', 14), ('Channel', 16), ('Customer', 18),
              ('City / Country', 14), ('Product ID', 16), ('Qty', 6), ('Order Value (₹)', 13),
              ('Payment', 15), ('Fulfilment', 12), ('Return Status', 14), ('Refund (₹)', 11),
              ('Refund Date', 12), ('Notes', 40)]
    for i, (name, width) in enumerate(o_cols, start=1):
        cell = ws_o.cell(row=1, column=i, value=name)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        ws_o.column_dimensions[get_column_letter(i)].width = width
    ws_o.freeze_panes = 'A2'
    ws_o.auto_filter.ref = 'A1:N400'

    if with_data:
        orders = [
            ('2026-08-02', 'ORD-8H2K', 'Website - COD', 'Customer 1', 'Hyderabad', '6343', 1, 7000,
             'COD - collected', 'Delivered', None, None, None, ''),
            ('2026-08-04', 'ORD-8J7Q', 'Website - COD', 'Customer 2', 'Secunderabad', '6988', 1, 999,
             'COD - collected', 'Delivered', None, None, None, ''),
            ('2026-08-05', 'ORD-8L4R', 'Website - Prepaid', 'Customer 3', 'Hyderabad', 'lc-zf-rose', 1, 35999,
             'Refunded', 'Delivered', 'Refunded', 35999, '2026-08-14',
             'Size exchange requested; piece restocked after inspection — see Stock Log RET-8L4R'),
            ('2026-08-07', 'ORD-8N1D', 'Website - COD', 'Customer 4', 'Warangal', 'HA4536', 1, 3800,
             'COD - collected', 'Delivered', None, None, None, ''),
            ('2026-08-09', 'ORD-8Q5S', 'In Store', 'Walk-in', 'Hyderabad', 'HA6320', 1, 4800,
             'Prepaid - paid', 'Delivered', None, None, None, 'Studio walk-in'),
            ('2026-08-11', 'ORD-8T3M', 'Website - COD', 'Customer 5', 'Vijayawada', 'HA3858', 1, 1299,
             'COD - to collect', 'Shipped', None, None, None, ''),
            ('2026-08-13', 'ORD-8V9A', 'WhatsApp / DM', 'Customer 6', 'Hyderabad', 'MI263', 1, 450,
             'Prepaid - paid', 'Delivered', None, None, None, 'Instagram DM enquiry'),
        ]
    else:
        orders = [('2026-08-01', 'ORD-0001', 'Website - COD', 'EXAMPLE — delete', 'Hyderabad',
                   'lc-example-1', 1, 1299, 'COD - to collect', 'Placed', None, None, None,
                   'EXAMPLE ROW — one row per order; fill Return/Refund columns only if a return happens')]
    for r, row in enumerate(orders, start=2):
        for i, v in enumerate(row, start=1):
            cell = ws_o.cell(row=r, column=i, value=v)
            cell.font = BASE
            cell.border = THIN
        ws_o.cell(row=r, column=1).number_format = 'yyyy-mm-dd'
        ws_o.cell(row=r, column=8).number_format = '"₹"#,##0'
        ws_o.cell(row=r, column=12).number_format = '"₹"#,##0'
        ws_o.cell(row=r, column=13).number_format = 'yyyy-mm-dd'
        if not with_data:
            for i in range(1, 15):
                ws_o.cell(row=r, column=i).fill = GREY_FILL
    for name, col in [('Channel', 'C'), ('Payment', 'I'), ('Fulfilment', 'J'), ('Return Status', 'K')]:
        dv = DataValidation(type='list', formula1=list_ranges[name], allow_blank=True)
        ws_o.add_data_validation(dv)
        dv.add(f'{col}2:{col}400')

    # ── Action Tracker sheet (the pending checklist) ─────────────────────────
    ws_t = wb.create_sheet('Action Tracker')
    t_cols = [('#', 4), ('Task', 62), ('Area', 12), ('Status', 12), ('Owner', 14),
              ('Done Date', 11), ('Notes', 46)]
    for i, (name, width) in enumerate(t_cols, start=1):
        cell = ws_t.cell(row=1, column=i, value=name)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        ws_t.column_dimensions[get_column_letter(i)].width = width
    ws_t.freeze_panes = 'A2'
    ws_t.auto_filter.ref = 'A1:G120'

    TASKS = [
        ('Rotate the Firebase service-account key (Firebase Console → Project settings → Service accounts → generate new, delete old)', 'Security', 'Pending', 'You', 'Current key was shared in chat — treat as exposed'),
        ('Regenerate the live Razorpay Key Secret; update RAZORPAY_KEY_SECRET in Vercel', 'Security', 'Pending', 'You', 'Old secret appeared in a screenshot'),
        ('Paste VITE_GSTIN (from the GST certificate) into Vercel and redeploy', 'Compliance', 'Pending', 'You', 'Until then invoices/footer omit the GSTIN by design'),
        ('Update VITE_BUSINESS_ADDRESS in Vercel to the full registered address (Gachibowli · 500046)', 'Compliance', 'Pending', 'You', 'Must match the GST certificate on invoices'),
        ('Add CRON_SECRET + GOOGLE_REVIEW_URL in Vercel to switch on the review-request emails', 'Marketing', 'Pending', 'You', 'Cron is deployed and waiting on these two values'),
        ('Create the free Microsoft Clarity project; paste VITE_CLARITY_PROJECT_ID into Vercel; redeploy', 'Marketing', 'Pending', 'You', 'Heatmaps + session recordings, consent-gated'),
        ('Connect Google Search Console (Domain property, DNS TXT) and submit sitemap.xml', 'SEO', 'Pending', 'You', '62 real URLs are live and waiting to be indexed'),
        ('Request indexing for home, Lehenga Cholis category, and 2–3 product pages (URL Inspection)', 'SEO', 'Pending', 'You', 'After Search Console verifies'),
        ('Google Business Profile: rewrite description (boutique, not fabric house), add secondary categories, In-store shopping, full address + PIN', 'Marketing', 'Pending', 'You', 'Draft description was provided in chat'),
        ('Google Business Profile: upload logo, cover photo, 15–25 real photos; set hours; add Products', 'Marketing', 'Pending', 'You', 'Profile strength currently 64%'),
        ('Collect the first 15–25 Google reviews (ask at handover + review-request email)', 'Marketing', 'Pending', 'You', 'Zero reviews today — the #1 local ranking factor'),
        ('CA: confirm GST rates (5% vs 18% over ₹2,500/piece) and HSN codes per category', 'Finance', 'Pending', 'CA', 'Then per-product rates get wired into checkout + invoices'),
        ('Replace illustrative buying prices in this workbook with real PO costs', 'Finance', 'Pending', 'You', 'Blue cells on the Catalogue sheet'),
        ('Reshoot 3 product photos: HA6378, MI263, RI5687 (warehouse snapshots today)', 'Catalogue', 'Pending', 'You', 'Flagged in Photo Quality column'),
        ('One live low-value Razorpay test payment after the key rotation', 'Finance', 'Pending', 'You', 'Proves the full prepaid path end to end'),
        ('Open Google Ads account with the new-advertiser spend-match credit (Search only)', 'Marketing', 'Blocked', 'You', 'Do AFTER reviews exist + GA4 conversions imported'),
        ('Set SENTRY_DSN in Vercel for server error alerting', 'Operations', 'Pending', 'You', 'Hooks are wired on all 12 API functions, currently inert'),
        ('If exporting: IEC registration + LUT filing (zero-rated exports without paying IGST)', 'Finance', 'Pending', 'CA', 'Only when international orders start'),
    ]
    for r, (task, area, status, owner, note) in enumerate(TASKS, start=2):
        for i, v in enumerate([r - 1, task, area, status, owner, None, note], start=1):
            cell = ws_t.cell(row=r, column=i, value=v)
            cell.font = BASE
            cell.border = THIN
            cell.alignment = Alignment(wrap_text=(i in (2, 7)), vertical='top')
        ws_t.row_dimensions[r].height = 30
    for name, col in [('Task Area', 'C'), ('Task Status', 'D')]:
        dv = DataValidation(type='list', formula1=list_ranges[name], allow_blank=True)
        ws_t.add_data_validation(dv)
        dv.add(f'{col}2:{col}120')
    green = PatternFill('solid', fgColor='D8EFDD')
    amber2 = PatternFill('solid', fgColor='FBEAC9')
    red2 = PatternFill('solid', fgColor='F8D3D0')
    ws_t.conditional_formatting.add('D2:D120', FormulaRule(formula=['$D2="Done"'], fill=green))
    ws_t.conditional_formatting.add('D2:D120', FormulaRule(formula=['$D2="Pending"'], fill=amber2))
    ws_t.conditional_formatting.add('D2:D120', FormulaRule(formula=['$D2="Blocked"'], fill=red2))

    # ── How To Use sheet ─────────────────────────────────────────────────────
    ws_h = wb.create_sheet('How To Use')
    ws_h.column_dimensions['A'].width = 4
    ws_h.column_dimensions['B'].width = 120
    lines = [
        ('TRESOR COUTURE — MASTER CATALOGUE (read this page first)', True),
        ('', False),
        ('WHAT EACH SHEET IS FOR', True),
        ('Catalogue = your products. One row per item you sell. This is where prices, stock and photos live.', False),
        ('Stock Log = a diary of stock. Every time stock moves — in or out — you add ONE line at the bottom. You never edit old lines.', False),
        ('Orders & Returns = your orders. One row per customer order. Refunds and return progress are tracked HERE, because a refund is about money, not stock.', False),
        ('Action Tracker = your to-do list. Filter the Status column to "Pending" to see everything left to do; set it to "Done" with a date when finished.', False),
        ('Lists = the options inside every dropdown. Add a new supplier or category here once and every dropdown learns it.', False),
        ('', False),
        ('THE STOCK LOG, EXPLAINED SIMPLY', True),
        ('Think of it as a bank passbook, but for stock instead of money. Stock coming in = deposit (+). Stock going out = withdrawal (−). '
         'The "Stock After" column is the balance after each line — watch it go 5 → 4 → 5 for a lehenga that was sold and then returned.', False),
        ('You only ever ADD a line at the bottom. The Catalogue compares this diary against the Stock Qty column and shows a red CHECK if they ever disagree — that is how you catch a movement nobody wrote down.', False),
        ('', False),
        ('WHICH LINE DO I ADD WHEN… (every scenario)', True),
        ('New stock arrives from a supplier  →  "Received (new stock)", qty +N, reference = PO number.', False),
        ('Someone buys on the website  →  "Sold - Online", qty −1, reference = order number. Also add the order on the Orders & Returns sheet.', False),
        ('Someone buys at the studio  →  "Sold - In Store", qty −1.', False),
        ('An export / international order ships  →  "Sold - Export", qty −1. (GST is zero-rated on exports — flag it to your CA.)', False),
        ('A customer returns an item and it is resellable  →  "Returned (back to stock)", qty +1. The REFUND goes on the Orders & Returns sheet, not here.', False),
        ('A customer returns an item and it is damaged  →  add NOTHING here (stock is not coming back); record the refund on Orders & Returns; optionally "Damaged / Write-off" if it was restocked first.', False),
        ('An item is damaged / lost in the studio  →  "Damaged / Write-off", qty −1.', False),
        ('You gift a piece or send a sample  →  "Sample / Gift", qty −1.', False),
        ('A physical count finds the sheet is wrong  →  "Count correction", qty +/− the difference, note why.', False),
        ('', False),
        ('REFUNDS IN ONE SENTENCE', True),
        ('Money out = Orders & Returns sheet (Refund column). Item back on the shelf = one "+1 Returned" line in the Stock Log. A refund with no restockable item touches ONLY the Orders sheet.', False),
        ('', False),
        ('HOW TO FILTER', True),
        ('Every sheet already has filter arrows in row 1. Examples: Action Tracker → Status arrow → tick "Pending" = everything still to do. '
         'Stock Log → Product ID arrow → one product = that product\'s full history. Orders & Returns → Return Status arrow = all open returns.', False),
        ('', False),
        ('CELL COLOURS', True),
        ('Blue text = numbers you typed (e.g. Buying Price). Black = calculated for you — do not overwrite (Margin %, Stock Status, Stock After, Log Match).', False),
        ('Red fill = out of stock, or the log disagrees with the stock column. Amber = at/below reorder level, or a Pending task. Green = a Done task.', False),
        ('', False),
        ('ASSUMPTIONS IN THE SAMPLE FILE', True),
        ('Buying prices are ILLUSTRATIVE (the website database stores no cost prices yet) — replace with real PO costs. '
         'GST is left at 5% with HSN blank pending your CA; garments above ₹2,500/piece likely attract 18%. '
         'The orders and stock movements are format examples using your real products and current stock levels. '
         'Generated from the live site database on 2026-08-16.', False),
    ]
    for r, (text, bold) in enumerate(lines, start=2):
        cell = ws_h[f'B{r}']
        cell.value = text
        cell.font = Font(name=ARIAL, size=11 if bold else 10, bold=bold)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        ws_h.row_dimensions[r].height = 15 if bold or not text else 30

    order = ['Catalogue', 'Stock Log', 'Orders & Returns', 'Action Tracker', 'How To Use', 'Lists']
    wb._sheets = [wb[t] for t in order]
    wb.save(path)
    print('wrote', path)


build(f'{SCRATCH}/tresor-catalogue-TEMPLATE.xlsx', with_data=False)
build(f'{SCRATCH}/tresor-catalogue-SAMPLE.xlsx', with_data=True)
